import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
from matplotlib.backends.backend_pdf import PdfPages
import seaborn as sns
from dotenv import load_dotenv
import logging
from pathlib import Path
from datetime import datetime
import traceback
import json
import sys

# Load environment variables
load_dotenv()

# Configure logging for console only
def setup_logging():
    """Setup console-only logging"""
    simple_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    # Setup root logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Console handler only
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(simple_formatter)
    logger.addHandler(console_handler)
    
    return logger

class ExcelToPDFConverter:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.conversion_results = []
        self.start_time = datetime.now()
        self.config = self._load_and_validate_config()
        
    def _load_and_validate_config(self):
        """Load and validate all configuration from environment variables"""
        config = {
            'excel_file_path': os.getenv('EXCEL_FILE_PATH_PDF'),
            'sheet_names': os.getenv('SHEET_NAMES_PDF', '').split(','),
            'base_pdf_path': os.getenv('SHEET_CONVERTED_PDF_PATH'),
            'max_retries': int(os.getenv('MAX_CONVERSION_RETRIES', '3')),
            'timeout_seconds': int(os.getenv('CONVERSION_TIMEOUT_SECONDS', '300')),
        }
        
        # Validate configuration
        errors = []
        
        if not config['excel_file_path']:
            errors.append("EXCEL_FILE_PATH_PDF environment variable not set")
        elif not os.path.exists(config['excel_file_path']):
            errors.append(f"Excel file not found: {config['excel_file_path']}")
        elif not config['excel_file_path'].lower().endswith(('.xlsx', '.xls')):
            errors.append(f"File is not a valid Excel file: {config['excel_file_path']}")
            
        if not config['sheet_names'] or config['sheet_names'] == ['']:
            errors.append("SHEET_NAMES_PDF environment variable not set or empty")
        else:
            # Clean sheet names
            config['sheet_names'] = [name.strip() for name in config['sheet_names'] if name.strip()]
            if not config['sheet_names']:
                errors.append("No valid sheet names found after cleaning")
                
        if not config['base_pdf_path']:
            errors.append("SHEET_CONVERTED_PDF_PATH environment variable not set")
        elif not os.path.exists(os.path.dirname(config['base_pdf_path'])):
            errors.append(f"Parent directory of PDF path does not exist: {os.path.dirname(config['base_pdf_path'])}")
            
        if errors:
            for error in errors:
                self.logger.error(f"Configuration error: {error}")
            raise ValueError(f"Configuration validation failed: {'; '.join(errors)}")
            
        self.logger.info("Configuration validated successfully")
        return config

    def setup_directories(self):
        """Create directories for each analyst if they don't exist"""
        directories = {}
        try:
            for analyst in self.config['sheet_names']:
                analyst_path = os.path.join(self.config['base_pdf_path'], analyst.title())
                Path(analyst_path).mkdir(parents=True, exist_ok=True)
                directories[analyst] = analyst_path
                self.logger.info(f"Directory ready: {analyst_path}")
                
                # Test write permissions
                test_file = os.path.join(analyst_path, "write_test.tmp")
                try:
                    with open(test_file, 'w') as f:
                        f.write("test")
                    os.remove(test_file)
                except Exception as e:
                    raise PermissionError(f"Cannot write to directory {analyst_path}: {str(e)}")
                    
        except Exception as e:
            self.logger.error(f"Error setting up directories: {str(e)}")
            raise
            
        return directories

    def get_available_sheets(self):
        """Get list of available sheets in the Excel file"""
        try:
            excel_file = pd.ExcelFile(self.config['excel_file_path'])
            available_sheets = excel_file.sheet_names
            self.logger.info(f"Available sheets in Excel file: {available_sheets}")
            return available_sheets
        except Exception as e:
            self.logger.error(f"Error reading Excel file structure: {str(e)}")
            raise

    def excel_to_pdf_simple(self, sheet_name, output_pdf_path):
        """Convert Excel sheet to PDF using openpyxl to preserve better formatting"""
        try:
            self.logger.debug(f"Starting simple PDF conversion for sheet: {sheet_name}")
            
            # Use openpyxl to read with better formatting preservation
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.config['excel_file_path'], data_only=True)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
                
            worksheet = workbook[sheet_name]
            
            # Get the actual used range (skip empty rows/columns)
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Find the actual data boundaries (skip empty leading rows/cols)
            min_row, min_col = 1, 1
            for row in range(1, max_row + 1):
                if any(worksheet.cell(row, col).value for col in range(1, max_col + 1)):
                    min_row = row
                    break
            
            for col in range(1, max_col + 1):
                if any(worksheet.cell(row, col).value for row in range(min_row, max_row + 1)):
                    min_col = col
                    break
            
            # Extract data preserving the layout
            data = []
            headers = []
            
            # Get headers from first row of data
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(min_row, col).value
                headers.append(str(cell_value) if cell_value is not None else "")
            
            # Get data rows, preserving empty cells as empty strings
            for row in range(min_row + 1, max_row + 1):
                row_data = []
                has_data = False
                for col in range(min_col, max_col + 1):
                    cell_value = worksheet.cell(row, col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value))
                        has_data = True
                    else:
                        row_data.append("")
                
                # Only add rows that have at least some data
                if has_data:
                    data.append(row_data)
            
            if not data:
                raise ValueError(f"No data found in sheet '{sheet_name}'")
            
            self.logger.debug(f"Sheet data loaded: {len(data)} rows, {len(headers)} columns")
            
            # Create PDF with better layout
            with PdfPages(output_pdf_path) as pdf:
                # Handle large datasets by splitting into pages
                max_rows_per_page = 40  # Reduced for better readability
                total_pages = (len(data) + max_rows_per_page - 1) // max_rows_per_page
                
                for page_num in range(total_pages):
                    start_idx = page_num * max_rows_per_page
                    end_idx = min((page_num + 1) * max_rows_per_page, len(data))
                    page_data = data[start_idx:end_idx]
                    
                    fig, ax = plt.subplots(figsize=(11.69, 8.27))  # A4 landscape
                    ax.axis('tight')
                    ax.axis('off')
                    
                    # Create table with proper data
                    table = ax.table(
                        cellText=page_data,
                        colLabels=headers,
                        cellLoc='center',
                        loc='center'
                    )
                    
                    # Improved styling
                    table.auto_set_font_size(False)
                    font_size = max(6, min(9, 100 // len(headers)))  # Dynamic font size based on columns
                    table.set_fontsize(font_size)
                    table.scale(1.0, 1.8)  # Better row height
                    
                    # Style header row
                    for i in range(len(headers)):
                        table[(0, i)].set_facecolor('#40466e')
                        table[(0, i)].set_text_props(weight='bold', color='white')
                    
                    # Alternate row colors for better readability
                    for i in range(1, len(page_data) + 1):
                        if i % 2 == 0:
                            for j in range(len(headers)):
                                table[(i, j)].set_facecolor('#f0f0f0')
                    
                    # Title with page info
                    title = f'Analysis Report - {sheet_name.title()}'
                    if total_pages > 1:
                        title += f' (Page {page_num + 1} of {total_pages})'
                    plt.title(title, fontsize=14, fontweight='bold', pad=20)
                    
                    pdf.savefig(fig, bbox_inches='tight', dpi=300)
                    plt.close()
            
            self.logger.info(f"Successfully created PDF: {output_pdf_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error in simple PDF conversion for {sheet_name}: {str(e)}")
            self.logger.debug(f"Full traceback: {traceback.format_exc()}")
            return False

    def excel_to_pdf_advanced(self, sheet_name, output_pdf_path):
        """Convert Excel sheet to PDF preserving formatting (requires win32com - Windows only)"""
        try:
            import win32com.client as win32
            from pythoncom import CoInitialize, CoUninitialize
            
            self.logger.debug(f"Starting advanced PDF conversion for sheet: {sheet_name}")
            
            # Initialize COM
            CoInitialize()
            
            try:
                # Convert to absolute paths
                excel_file_path = os.path.abspath(self.config['excel_file_path'])
                output_pdf_path = os.path.abspath(output_pdf_path)
                
                # Open Excel application
                excel_app = win32.Dispatch("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                excel_app.ScreenUpdating = False
                
                # Open workbook
                workbook = excel_app.Workbooks.Open(excel_file_path, ReadOnly=True)
                
                # Verify sheet exists
                sheet_names = [ws.Name for ws in workbook.Worksheets]
                if sheet_name not in sheet_names:
                    raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
                
                # Select the specific sheet
                worksheet = workbook.Sheets(sheet_name)
                
                # Check if sheet has data
                used_range = worksheet.UsedRange
                if used_range is None:
                    raise ValueError(f"Sheet '{sheet_name}' has no data")
                
                # Export to PDF with simplified parameters (avoiding compatibility issues)
                worksheet.ExportAsFixedFormat(
                    Type=0,  # PDF format
                    Filename=output_pdf_path,
                    Quality=0,  # 0 = minimum size, 1 = maximum quality
                    IgnorePrintAreas=False,
                    OpenAfterPublish=False
                )
                
                # Close workbook and quit Excel
                workbook.Close(False)
                excel_app.Quit()
                
            finally:
                CoUninitialize()
            
            self.logger.info(f"Successfully created PDF with formatting: {output_pdf_path}")
            return True
            
        except ImportError:
            self.logger.warning("win32com not available. Using simple PDF conversion method.")
            return self.excel_to_pdf_simple(sheet_name, output_pdf_path)
        except Exception as e:
            self.logger.error(f"Error with advanced PDF conversion for {sheet_name}: {str(e)}")
            self.logger.debug(f"Full traceback: {traceback.format_exc()}")
            # Fallback to simple method
            return self.excel_to_pdf_simple(sheet_name, output_pdf_path)

    def convert_sheet_with_retry(self, sheet_name, output_pdf_path):
        """Convert a single sheet with retry logic"""
        max_retries = self.config['max_retries']
        
        for attempt in range(1, max_retries + 1):
            try:
                self.logger.info(f"Converting sheet '{sheet_name}' (attempt {attempt}/{max_retries})")
                
                # Try advanced method first, fallback to simple
                success = self.excel_to_pdf_advanced(sheet_name, output_pdf_path)
                
                if success and os.path.exists(output_pdf_path) and os.path.getsize(output_pdf_path) > 0:
                    file_size = os.path.getsize(output_pdf_path)
                    self.logger.info(f"Conversion successful. PDF size: {file_size:,} bytes")
                    return True, None
                else:
                    raise Exception("PDF file was not created or is empty")
                    
            except Exception as e:
                error_msg = str(e)
                self.logger.warning(f"Attempt {attempt} failed for sheet '{sheet_name}': {error_msg}")
                
                if attempt == max_retries:
                    self.logger.error(f"All {max_retries} attempts failed for sheet '{sheet_name}'")
                    return False, error_msg
                else:
                    self.logger.info(f"Retrying in 2 seconds...")
                    import time
                    time.sleep(2)
        
        return False, "Maximum retry attempts exceeded"

    def generate_report(self, directories):
        """Generate comprehensive success/failure report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = "reports"
        Path(report_dir).mkdir(exist_ok=True)
        
        # Text report
        report_file = os.path.join(report_dir, f"conversion_report_{timestamp}.txt")
        
        # JSON report for programmatic access
        json_report_file = os.path.join(report_dir, f"conversion_report_{timestamp}.json")
        
        success_count = sum(1 for result in self.conversion_results if result['success'])
        total_count = len(self.conversion_results)
        duration = datetime.now() - self.start_time
        
        # Generate text report
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("EXCEL TO PDF CONVERSION REPORT\n")
            f.write("=" * 80 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Duration: {duration}\n")
            f.write(f"Source File: {self.config['excel_file_path']}\n")
            f.write(f"Output Base Path: {self.config['base_pdf_path']}\n")
            f.write(f"Success Rate: {success_count}/{total_count} ({(success_count/total_count*100):.1f}%)\n")
            f.write("\n")
            
            if success_count > 0:
                f.write("SUCCESSFUL CONVERSIONS:\n")
                f.write("-" * 40 + "\n")
                for result in self.conversion_results:
                    if result['success']:
                        f.write(f"✓ {result['sheet_name']}\n")
                        f.write(f"  Output: {result['output_path']}\n")
                        f.write(f"  Duration: {result['duration']}\n")
                        if result.get('file_size'):
                            f.write(f"  File Size: {result['file_size']:,} bytes\n")
                        f.write("\n")
            
            if success_count < total_count:
                f.write("FAILED CONVERSIONS:\n")
                f.write("-" * 40 + "\n")
                for result in self.conversion_results:
                    if not result['success']:
                        f.write(f"✗ {result['sheet_name']}\n")
                        f.write(f"  Error: {result['error']}\n")
                        f.write(f"  Attempted Output: {result['output_path']}\n")
                        f.write("\n")
            
            # Configuration summary
            f.write("CONFIGURATION:\n")
            f.write("-" * 40 + "\n")
            f.write(f"Max Retries: {self.config['max_retries']}\n")
            f.write(f"Timeout: {self.config['timeout_seconds']} seconds\n")
            f.write(f"Requested Sheets: {', '.join(self.config['sheet_names'])}\n")
        
        # Generate JSON report
        json_data = {
            'timestamp': datetime.now().isoformat(),
            'duration_seconds': duration.total_seconds(),
            'source_file': self.config['excel_file_path'],
            'output_base_path': self.config['base_pdf_path'],
            'total_sheets': total_count,
            'successful_conversions': success_count,
            'success_rate': success_count / total_count if total_count > 0 else 0,
            'configuration': self.config,
            'results': self.conversion_results
        }
        
        with open(json_report_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, default=str)
        
        self.logger.info(f"Reports generated:")
        self.logger.info(f"  Text report: {report_file}")
        self.logger.info(f"  JSON report: {json_report_file}")
        
        return report_file, json_report_file

    def run(self):
        """Main execution method"""
        try:
            self.logger.info("Starting Excel to PDF conversion process")
            self.logger.info(f"Source: {self.config['excel_file_path']}")
            self.logger.info(f"Sheets to convert: {self.config['sheet_names']}")
            self.logger.info(f"Output base: {self.config['base_pdf_path']}")
            
            # Setup directories
            directories = self.setup_directories()
            
            # Get available sheets
            available_sheets = self.get_available_sheets()
            
            # Process each requested sheet
            for sheet_name in self.config['sheet_names']:
                start_time = datetime.now()
                
                # Check if sheet exists
                if sheet_name not in available_sheets:
                    error_msg = f"Sheet '{sheet_name}' not found in Excel file. Available: {available_sheets}"
                    self.logger.warning(error_msg)
                    self.conversion_results.append({
                        'sheet_name': sheet_name,
                        'success': False,
                        'error': error_msg,
                        'output_path': 'N/A',
                        'duration': datetime.now() - start_time
                    })
                    continue
                
                # Create output path with timestamp
                analyst_folder = directories[sheet_name]
                timestamp = datetime.now().strftime("%m.%d.%Y_%H%M")
                pdf_filename = f"{sheet_name}_analysis_report_{timestamp}.pdf"
                output_pdf_path = os.path.join(analyst_folder, pdf_filename)
                
                # Attempt conversion
                success, error = self.convert_sheet_with_retry(sheet_name, output_pdf_path)
                
                # Record result
                result = {
                    'sheet_name': sheet_name,
                    'success': success,
                    'output_path': output_pdf_path,
                    'duration': datetime.now() - start_time
                }
                
                if success:
                    result['file_size'] = os.path.getsize(output_pdf_path)
                else:
                    result['error'] = error
                
                self.conversion_results.append(result)
            
            # Generate reports
            text_report, json_report = self.generate_report(directories)
            
            # Final summary
            success_count = sum(1 for r in self.conversion_results if r['success'])
            total_count = len(self.conversion_results)
            
            self.logger.info("=" * 60)
            self.logger.info(f"CONVERSION COMPLETE: {success_count}/{total_count} successful")
            
            if success_count == total_count:
                self.logger.info("🎉 All conversions successful!")
                return 0  # Success exit code
            elif success_count > 0:
                self.logger.warning(f"⚠️  Partial success: {total_count - success_count} failed")
                return 1  # Partial failure exit code
            else:
                self.logger.error("❌ All conversions failed!")
                return 2  # Complete failure exit code
                
        except Exception as e:
            self.logger.error(f"Critical error in main process: {str(e)}")
            self.logger.debug(f"Full traceback: {traceback.format_exc()}")
            return 3  # Critical error exit code

def main():
    """Entry point with comprehensive error handling"""
    logger = setup_logging()
    
    try:
        logger.info("Excel to PDF Converter Starting...")
        
        converter = ExcelToPDFConverter()
        exit_code = converter.run()
        
        logger.info(f"Process completed with exit code: {exit_code}")
        return exit_code
        
    except KeyboardInterrupt:
        logger.warning("Process interrupted by user")
        return 130  # Standard exit code for Ctrl+C
    except Exception as e:
        logger.error(f"Unhandled exception: {str(e)}")
        return 4  # Unhandled error exit code
    finally:
        logging.shutdown()

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)